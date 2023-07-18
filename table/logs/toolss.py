import re
import openpyxl
from openpyxl.styles import PatternFill , Font
import sys
# Create a new Excel workbook
wb = openpyxl.Workbook()

# Initialize variables
current_proxy_id = None
sheet = None  # Initialize sheet variable here

# Read the log file
if __name__ == '__main__':
    file_name = sys.argv[1]  # Retrieve the file name from command-line argument
    with open(file_name, 'r') as file:
        lines = file.readlines()

# Initialize row counter and MO counter
    row = 2
    mo_counter = 1

# Process each line in the log file
    for line in lines:
        line = line.strip()

    # Check for Proxy Id pattern
        if line.startswith("Proxy Id"):
            current_proxy_id = line.split()[-1]
            sheet = wb.create_sheet(title=f"Proxy {current_proxy_id}")


        # Set column headings in light blue color
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)
            sheet["A1"] = "MO"
            sheet["B1"] = "Attribute"
            sheet["C1"] = "Value"
            for cell in ["A1", "B1", "C1"]:
                sheet[cell].fill = header_fill
                sheet[cell].font = header_font

            # Initialize row counter and MO counter for each sheet
            row = 2
            mo_counter = 1
            continue

    # Extract the attribute and value
        match = re.match(r'([^ ]+)\s+(.*)', line)
        if match and sheet is not None:
            attribute = match.group(1)
            value = match.group(2)

        # Check if the line has the special format ">>> attribute = value"
            special_match = re.match(r'>>>(\s+)([^=]+)=(.*)', line)
            if special_match:
                attribute = special_match.group(2).strip()
                value = special_match.group(3).strip()

        # Write the attribute and value to the Excel sheet
            sheet.cell(row=row, column=1).value = mo_counter  # MO
            sheet.cell(row=row, column=2).value = attribute
            sheet.cell(row=row, column=3).value = value
            row += 1
            mo_counter += 1

# Save the Excel workbook
wb.save('sorted_data.xlsx')
