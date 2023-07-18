from django.shortcuts import render, redirect, HttpResponseRedirect
from django.template import loader
from myapp.models import companydata
from django.http import HttpResponse
from myapp.models import infos
from django.contrib import messages
from myapp.models import UserList
from myapp.models import clientname
from myapp.models import markettime
from myapp.models import marketdata
from myapp.models import operator
from myapp.models import activitylists
from myapp.models import oems
from django.contrib.auth import logout
from django import forms 



import re
import openpyxl
from openpyxl.styles import PatternFill, Font
from django.http import HttpResponse
from django.views.decorators.http import require_POST

from .forms import FileUploadForm
from django.shortcuts import render
from django.http import HttpResponse
from .forms import FileUploadForm
from .models import Uploaded_file




def school(request):

    return render(request, 'school.html')





def process_excel(request):
    if request.method == 'POST':
        input_file = request.FILES['inputFile']
        fa_codes_str = request.POST['faCodes']
        fa_codes = [int(code) for code in fa_codes_str.split(",")]

        excel_data = pd.read_excel(input_file, sheet_name=None)
        cell_info_data = excel_data['Cell info']
        filtered_data = cell_info_data[cell_info_data['FA Code'].isin(fa_codes)]
        filtered_df = pd.DataFrame(filtered_data)

        output_wb = Workbook()
        output_ws = output_wb.active

        for col_num, column_name in enumerate(filtered_df.columns, 1):
            output_ws.cell(row=1, column=col_num, value=column_name)

        for row in filtered_df.itertuples(index=False, name=None):
            output_ws.append(row)

        for col_num in range(1, filtered_df.shape[1] + 1):
            output_ws.cell(row=1, column=col_num).fill = PatternFill(start_color="7EC0EE", fill_type="solid")

        output_file = os.path.join(settings.MEDIA_ROOT, 'output.xlsx')
        output_wb.save(output_file)

        output_file_url = settings.MEDIA_URL + 'output.xlsx'
        return render(request, 'process_excel.html', {'output_file_url': output_file_url})

    return render(request, 'process_excel.html')



# --------------------------------------------------EDP--------------------------------------------

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse

def process_form(request):
    if request.method == 'POST':
        # Get the uploaded file and site_usid values from the form
        uploaded_file = request.FILES['file']
        site_usids = request.POST['site_usids'].split(',')

        # Save the uploaded file
        fs = FileSystemStorage(location=settings.MEDIA_ROOT)
        input_file = fs.save(uploaded_file.name, uploaded_file)

        # Read the input Excel file
        df = pd.read_excel(fs.path(input_file))

        # Filter rows based on the user-provided site_usid values
        filtered_df = df[df['site_usid'].astype(str).isin(site_usids)]

        # Make a copy of the filtered DataFrame
        filtered_df = filtered_df.copy()

        # Fill empty cells or columns with 'NA'
        filtered_df.fillna('NA', inplace=True)

        # Create a new Excel workbook
        output_file = 'output.xlsx'
        output_path = os.path.join(settings.MEDIA_ROOT, output_file)

        # Create a workbook and worksheet
        workbook = Workbook()
        worksheet = workbook.active

        # Write the column names to the worksheet
        worksheet.append(filtered_df.columns.tolist())

        # Write the filtered DataFrame rows to the worksheet
        for _, row in filtered_df.iterrows():
            worksheet.append(row.tolist())

        # Apply the sky blue fill color to the first row (including column names)
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill(fill_type='solid', fgColor='00BFFF')

        # Save the workbook
        workbook.save(output_path)

        # Provide the download URL for the template
        download_url = fs.url(output_file)

        return render(request, 'form.html', {'download_url': download_url})

    return render(request, 'form.html')



import re
import openpyxl
from openpyxl.styles import PatternFill, Font
from django.shortcuts import render
from django.http import HttpResponse

def process_log_file(file_path):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Initialize variables
    current_proxy_id = None
    sheet = None  # Initialize sheet variable here

    # Read the log file
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Initialize row counter and MO counter
    row = 2
    mo_counter = 1

    # Process each line in the log file
    for line in lines:
        line = line.strip()

        # Check for Proxy Id pattern
        if line.startswith("Proxy Id"):
            current_proxy_id = line.split()[-1]
            sheet = wb.create_sheet(title=f"Proxy {current_proxy_id}")

            # Set column headings in light blue color
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)
            sheet["A1"] = "MO"
            sheet["B1"] = "Attribute"
            sheet["C1"] = "Value"
            for cell in ["A1", "B1", "C1"]:
                sheet[cell].fill = header_fill
                sheet[cell].font = header_font

            # Initialize row counter and MO counter for each sheet
            row = 2
            mo_counter = 1
            continue

        # Extract the attribute and value
        match = re.match(r'([^ ]+)\s+(.*)', line)
        if match and sheet is not None:
            attribute = match.group(1)
            value = match.group(2)

            # Check if the line has the special format ">>> attribute = value"
            special_match = re.match(r'>>>(\s+)([^=]+)=(.*)', line)
            if special_match:
                attribute = special_match.group(2).strip()
                value = special_match.group(3).strip()

            # Write the attribute and value to the Excel sheet
            sheet.cell(row=row, column=1).value = mo_counter  # MO
            sheet.cell(row=row, column=2).value = attribute
            sheet.cell(row=row, column=3).value = value
            row += 1
            mo_counter += 1

    # Save the Excel workbook
    wb.save('sorted_data.xlsx')

def upload_file(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if file:
            file_path = 'media/' + file.name
            with open(file_path, 'wb') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            process_log_file(file_path)
            return render(request, 'lgfile.html', {'file_processed': True})

    return render(request, 'lgfile.html', {'file_processed': False})

def download_file(request):
    file_path = 'sorted_data.xlsx'
    with open(file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sorted_data.xlsx'
        return response




# ----------------------------------------------------------------------------------------------------------------------------------------
#  create
def index1(request):

    return render(request, 'index1.html')


def successful(request):

    return render(request, 'successful.html')

 

def activitylist(request):
    activitylist = activitylists.objects.all()
    
     
    context = {
        'activitylist': activitylist, 
             
      
    }
    return render(request, 'activitylist.html',context)

 
def signout(request):
    logout(request)
    messages.error(request, "Logout Successfully")
    return redirect('/')


def login_page(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')
        if infos.objects.filter(email=email, password=password).exists():
            # Store the email in the session
            request.session['email'] = email
            # messages.success(request, "Login Successfully!")
            # Retrieve the 'infos' object
            info = infos.objects.get(email=email, password=password)
            
              
            return render(request, 'account_settings.html', {'email': email, 'password': password, 'infos': info})
        else:
            messages.error(request, "Invalid Credentials!")
            return redirect('/login_page')

    

    return render(request, 'login.html',)
 



def signup(request):
    if request.method == "POST":
        firstname = request.POST.get('firstname')
        secondname = request.POST.get('secondname')
        company_id = request.POST.get('companyid')
        roleid = request.POST.get('role')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        password = request.POST.get('password')
     
        
        
        if infos.objects.filter(email=email).exists():
            messages.error(request, "Email already exists!")
            return redirect('/signup')
        
        if infos.objects.filter(phone=phone).exists():
            messages.error(request, "Phone  already exists!")
            return redirect('/signup')
        
          
        

        try:
            role = UserList.objects.get(roleid=roleid)
            company = companydata.objects.get(companyid=company_id)
        except UserList.DoesNotExist:
            messages.error(request, "Invalid role.")
            return redirect('signup')
        except companydata.DoesNotExist:
            messages.error(request, "Invalid company ID.")
            return redirect('signup')

        infos_instance = infos(
            firstname=firstname,
            secondname=secondname,
            companyid=company,
            role=role,
            email=email,
            phone=phone,
            password=password
          
            
        )

        infos_instance.save()
        messages.success(request, "You have been successfully signed up.")
        return redirect("/signup")

    context = {
        'role': UserList.objects.all(),
        
        'companies': companydata.objects.values('companyid', 'company')  # Query both company ID and company name
    }
    return render(request, 'signup.html', context)




def datatable(request):
    datatable = infos.objects.all()
    role_data = UserList.objects.filter(roleid__in=[info.role.roleid for info in datatable])
    company_data = companydata.objects.filter(companyid__in=[info.companyid.companyid for info in datatable])

    context = {
        'datatable': datatable,
        'role_data': role_data,
    }
    context = {
        'datatable': datatable,
        'role_data': company_data,
    }

    return render(request, 'datatable.html', context)

def createactivity(request):
    market_state_choices = activitylists.STATES
    market_state_choices2 = activitylists.sow_type


    context = {
                'market_state_choices2': market_state_choices2,

        'market_state_choices': market_state_choices,
        'markettime': markettime.objects.all(),
        'operatorid': operator.objects.values('operatorid', 'operatorname'),
        'client': clientname.objects.values('clientid', 'cname')
    }

    return render(request, 'create_activity.html', context)



import pandas as pd










import re
import openpyxl
from openpyxl.styles import PatternFill, Font
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.http import HttpResponse














def forgot(request):
    if request.method == "POST":
        email = request.POST.get('email')
        if infos.objects.filter(email=email).exists():
            messages.success(request, "Email sent successfullty!")
            return redirect('/emailreset')

        else:
            messages.error(request, "Email is Incorrect!")
            return redirect('/forgot')
    return render(request, 'forgot.html')


def emailreset(request):
    return render(request, 'emailreset.html')

def adduser(request):
    return render(request, 'adduser.html')

def add_market(request):
     if request.method == "POST":
        clientid = request.POST.get('clientid')
        operator_id = request.POST.get('operatorid')
        oemid = request.POST.get('oemid')
        marketname = request.POST.get('marketname')
        timeid = request.POST.get('timeid')
        Specific_Market_Guidline = request.POST.get('Specific_Market_Guidline')
        Ciq = request.FILES['Ciq']
        Call_test_infos = request.FILES['Call_test_info']
        Call_test_files = request.FILES['Call_test_files']
        Site_accesses = request.FILES['Site_access']
        Guideliness = request.FILES['Guidelines']
        additional_Guideliness = request.FILES['additional_Guidelines']
        
       



        try:
            cname = clientname.objects.get(clientid=clientid)
            operatorname = operator.objects.get(operatorid=operator_id)
            mtime = markettime.objects.get(timeid=timeid)
            oem_name = oems.objects.get(oemid=oemid)
        
        except clientname.DoesNotExist:
            messages.error(request, "Invalid clientname.")
            return redirect('add_market')
        except companydata.DoesNotExist:
            messages.error(request, "Invalid operator.")
            return redirect('add_market')
        except markettime.DoesNotExist:
            messages.error(request, "Invalid markettime.")
            return redirect('add_market')
        except oems.DoesNotExist:
            messages.error(request, "Invalid oem.")
            return redirect('add_market')

        marketdata_obj = marketdata(
            clientid=cname,
            operatorid=operatorname,
            oemid=oem_name,
            marketname=marketname,
            timeid=mtime,
            Specific_Market_Guidline=Specific_Market_Guidline,
            Ciq=Ciq,
            Call_test_info=Call_test_infos,
            Call_test_files=Call_test_files,
            Site_access=Site_accesses,
            Guidelines=Guideliness,
            additional_Guidelines=additional_Guideliness         
            
        )
            
        marketdata_obj.save()
        messages.success(request, "You have been successfully signed up.")
        return redirect("/add_market")
    
  


    
     context = {
        'clientid': clientname.objects.all(),
        'timeid': markettime.objects.all(),
        'operatorid': operator.objects.all(),
        'oemid': oems.objects.all(),
        
       
     }
    
     return render(request, 'add_market.html', context)

# def add_market(request):
#     if request.method == "POST":
#         clientids = request.POST.getlist('clientid')
#         operator_ids = request.POST.getlist('operatorid')
#         oems = request.POST.getlist('oem')
#         marketnames = request.POST.getlist('marketname')
#         timeids = request.POST.getlist('timeid')
#         Specific_Market_Guidlines = request.POST.getlist('Specific_Market_Guidline')
#         Ciqs = request.FILES.getlist('Ciq')
#         Call_test_infos = request.FILES.getlist('Call_test_info')
#         Call_test_files = request.FILES.getlist('Call_test_files')
#         Site_accesses = request.FILES.getlist('Site_access')
#         Guideliness = request.FILES.getlist('Guidelines')
#         additional_Guideliness = request.FILES.getlist('additional_Guidelines')

#         try:
#             cnames = [clientname.objects.get(clientid=clientid) for clientid in clientids]
#             operatornames = [operator.objects.get(operatorid=operator_id) for operator_id in operator_ids]
#             mtimes = [markettime.objects.get(timeid=timeid) for timeid in timeids]
#         except clientname.DoesNotExist:
#             messages.error(request, "Invalid clientname.")
#             return redirect('add_market')
#         except companydata.DoesNotExist:
#             messages.error(request, "Invalid operator.")
#             return redirect('add_market')
#         except markettime.DoesNotExist:
#             messages.error(request, "Invalid markettime.")
#             return redirect('add_market')

#         if not clientids:
#             marketdata_objs = []
#         else:
#             marketdata_objs = []
#             for i in range(len(clientids)):
#                 marketdata_objs.append(
#                     marketdata(
#                         clientid=cnames[i],
#                         operatorid=operatornames[clientids],
#                         oem=oems[i],
#                         marketname=marketnames[clientids],
#                         timeid=mtimes[clientids],
#                         Specific_Market_Guidline=Specific_Market_Guidlines[clientids],
#                         Ciq=Ciqs[i],
#                         Call_test_info=Call_test_infos[i],
#                         Call_test_files=Call_test_files[i],
#                         Site_access=Site_accesses[i],
#                         Guidelines=Guideliness[i],
#                         additional_Guidelines=additional_Guideliness[i]
#                     )
#                 )

#         marketdata_objs.save()
#         messages.success(request, "Market add Successfully.")

#         return redirect("/add_market")

#     context = {
#         'clientid': clientname.objects.all(),
#         'timeid': markettime.objects.all(),
#         'operatorid': operator.objects.all(),
#     }

#     return render(request, 'add_market.html', context)



def alarm(request):
    return render(request, 'alarm.html')



def account_settings(request):
    if request.method == "POST":
        # Check the form submission URL.
        if request.path != '/account_settings':
            messages.error(request, 'Invalid form submission URL.')
            return render(request, 'account_settings.html')

        # Retrieve the existing infos object for the logged-in user.
        try:
            email = request.session.get('email')  # Assuming 'email' is stored in the session after successful login
            if not email:
                raise ValueError('Email not found in the session.')

            info = infos.objects.get(email=email)  # Replace 'email' with the actual field name for user identification
        except infos.DoesNotExist:
            messages.error(request, 'User information not found.')
            return render(request, 'account_settings.html')
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, 'account_settings.html')

        print(f"Retrieved email: {email}")

        # Update the infos object.
        info.firstname = request.POST.get('firstname')
        info.secondname = request.POST.get('secondname')
        UserList.role = request.POST.get('role')
        companydata.companyid = request.POST.get('companyname')
        info.email = request.POST.get('email')
        info.phone = request.POST.get('phone')


        if request.FILES.get('image'):
            info.image = request.FILES.get('image')
            
       
        # Save the infos object to the database.
        info.save()
        UserList.save(info)
        companydata.save(info)
       
        
        

        # Redirect the user to the accountsettings page.
        return redirect('/account_settings')
     

    context = {
        'role': UserList.objects.all(),
        
        'companies': companydata.objects.values('companyid', 'company')  # Query both company ID and company name
    }
    

    return render(request, 'account_settings.html', context)

        

def dashboard(request):
    return render(request, 'dashboard.html',)

def practice(request):
    market2 = marketdata.objects.all()
    client_name = clientname.objects.filter(clientid__in=[info.clientid.id for info in market2])
    time_id = markettime.objects.filter(timeid__in=[info.timeid.id for info in market2])
    operator_id = operator.objects.filter(operatorid__in=[info.operatorid.id for info in market2])

    context = {
        'market2': market2,
        'clientname': client_name,
    }
    context = {
        'market2': market2,
        'timeid': time_id,
    }
    context = {
        'market2': market2,
        'operatorid': operator_id,
    }

    return render(request, 'practice.html', context)

def market2(request):  
     if request.method == 'POST':
        marketdata_id = request.POST.get('marketdata_id')
        if marketdata_id:
            try:
                
                
                marketdata_obj = marketdata.objects.get(id=marketdata_id)
                # Fetch the data for the specific marketdata_id from the MarketData model
                foreign_key_values = operator.objects.values_list('id', 'operatorname')
                markettime_values = markettime.objects.values_list('id', 'mtime')
                client_values = clientname.objects.values_list('id', 'cname')
                oem_values = oems.objects.values_list('id', 'oem_name')

                # Render the template with the marketdata object and dropdown values
                return render(request, 'edit_market.html', {'marketdata': marketdata_obj, 'dropdown_values': foreign_key_values, 'markettime_values': markettime_values, 'client_values': client_values,  'oem_values': oem_values})
            except marketdata.DoesNotExist:
                # Handle the case when marketdata_id is not found
                return HttpResponse('Market data does not exist')
        else:
            # Handle the case when marketdata_id is not provided
            return redirect('/market2') 

    # Handle the GET request to display the edit form
     data = marketdata.objects.all()

    # Get the default marketdata object for the first row
     default_marketdata_obj = data.first()
    

    # Get the values for the dropdown from the foreign key relationship
     foreign_key_values = operator.objects.values_list('id', 'operatorname')
     markettime_values = markettime.objects.values_list('id', 'mtime')
     client_values = clientname.objects.values_list('id', 'cname')
     oem_values = oems.objects.values_list('id', 'oem_name')
     
  

     return render(request, 'market2.html', {'market2': data,'marketdata': default_marketdata_obj, 'dropdown_values': foreign_key_values, 'markettime_values': markettime_values, 'client_values': client_values, 'oem_values': oem_values})


def delete(request):
    # Get the table data object from the database
    table_data = marketdata.objects.get(id=request.GET.get('id'))

    # Delete the table data object
    table_data.delete()

    # Redirect the user back to the index page
    return HttpResponseRedirect('/table/')


def userlist(request):
    if request.method == 'POST':
        emailinput = request.POST.get('search_name')
        print(f"The following email has been deleted: {emailinput}")

        try:
                
                row = infos.objects.get(email=emailinput)
                deleted_email = row.email  # Get the email value before deletion

                row.delete()
                # messages.error(request, "Invalid Credentials!")
                messages.success(request, f"Market data added successfully. Clone count: {deleted_email}") # Display email value in console
                return redirect('/userlist') 

        except infos.DoesNotExist:
                print(f"The following email has NOT been deleted: {emailinput}")  # Display email value in console
                return redirect('/userlist') 
    # Retrieve the user list and related role data
    datatable = infos.objects.all()
    role_data = UserList.objects.filter(roleid__in=[info.role.roleid for info in datatable])
    company_data = companydata.objects.filter(companyid__in=[info.companyid.companyid for info in datatable])

    context = {
        'datatable': datatable,
        'role_data': role_data,
    }
    context = {
        'datatable': datatable,
        'role_data': company_data,
    }

    return render(request, 'userlist.html', context)




def rolelist(request):
    if request.method == 'POST':
        roleid = request.POST.get('roleid')
        rolename = request.POST.get('rolename')
        
        # Create a new UserList object and save it to the database
        new_role = UserList(roleid=roleid, role=rolename)
        new_role.save()
        
        # Redirect to the rolelist page or any other desired page
        messages.success(request, "Role data added successfully")
        return redirect('rolelist')  
    
    # Retrieve data for the rolelist
    data = UserList.objects.all()
    
    context = {
        'rolelist': data
    }
    return render(request, 'rolelist.html', context)


def companylist(request):
    if request.method == 'POST':
        companyid = request.POST.get('companyid')
        company = request.POST.get('company')
        
        # Create a new UserList object and save it to the database
        new_company = companydata(companyid=companyid, company=company)
        new_company.save()
        
        # Redirect to the rolelist page or any other desired page
        messages.success(request, "Company added successfully")
        return redirect('/companylist')  
    
    # Retrieve data for the rolelist
    data = companydata.objects.all()
    
    context = {
        'company': data
    }
    return render(request, 'companylist.html', context)

def edit_market(request): 
    if request.method == "POST":
        # Check the form submission URL.
        if request.path != '/edit_market':
            messages.error(request, 'Invalid form submission URL.')
            return render(request, 'edit_market.html')

        # Retrieve the existing infos object for the logged-in user.
        try:
            id = request.session.get('id')  # Assuming 'email' is stored in the session after successful login
            if not id:
                raise ValueError('Email not found in the session.')

            info = marketdata.objects.get(id=id)  # Replace 'email' with the actual field name for user identification
        except marketdata.DoesNotExist:
            messages.error(request, 'User information not found.')
            return render(request, 'edit_market.html')
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, 'edit_market.html')
 
        print(f"Retrieved email: {id}")

        # Update the infos object.
        clientname.clientid = request.POST.get('clientid')
        operator.operatorid = request.POST.get('operatorid')
        oems.oemid = request.POST.get('oemid')
        info.marketname = request.POST.get('marketname')
        markettime.timeid = request.POST.get('timeid')
        info.Specific_Market_Guidline = request.POST.get('Specific_Market_Guidline')
        info.Ciq = request.POST.get('Ciq')
        info.Call_test_info = request.POST.get('Call_test_info')
        info.Call_test_files = request.POST.get('Call_test_files')
        info.Site_access = request.POST.get('Site_access')
        info.Guidelines = request.POST.get('Guidelines')
        info.additional_Guidelines = request.POST.get('additional_Guidelines')

            
       
        # Save the infos object to the database.
        info.save()
        clientname.save(info)
        operator.save(info)
        oems.save(info)
        markettime.save(info)
       
        
    
        # Redirect the user to the accountsettings page.
        return redirect('/edit_market')
     

    context = {
        'clientid': clientname.objects.all(),
        
        'operatorid': operator.objects.all(),
        'oemid': oems.objects.all(),
        'mtime': markettime.objects.all(),
        
        
    }
    

    return render(request, 'edit_market.html', context)

    
   

    
 



    